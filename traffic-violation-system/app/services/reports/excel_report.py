import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from app.services.evidence.evidence_service import evidence_service

class ExcelReportGenerator:
    @staticmethod
    def generate(report_id: int, name: str) -> str:
        """
        Creates a beautifully formatted Excel report using real database / evidence locker records.
        """
        reports_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "reports"))
        os.makedirs(reports_dir, exist_ok=True)
        filepath = os.path.join(reports_dir, f"{name}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Violations Summary"

        # Headers styling
        title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        normal_font = Font(name="Calibri", size=11)
        
        blue_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
        purple_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")

        # 1. Title Block
        ws.merge_cells("A1:G1")
        ws["A1"] = "AURA SMART MONITOR SYSTEM - Operations & Infractions Summary Report"
        ws["A1"].font = title_font
        ws["A1"].fill = blue_fill
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 30

        # 2. Metadata Block
        ws["A3"] = "Report ID"
        ws["B3"] = report_id
        ws["A4"] = "Generated At"
        ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["A5"] = "Scope"
        ws["B5"] = "Daily Control Center Summary"

        for row in range(3, 6):
            ws.cell(row=row, column=1).font = Font(bold=True)

        # 3. Table Headers
        headers = ["Evidence ID", "Violation ID", "Vehicle ID", "Plate Number", "Violation Type", "Confidence", "Camera"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_idx)
            cell.value = h
            cell.font = header_font
            cell.fill = purple_fill
            cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[7].height = 20

        # Load real data
        all_evidence = evidence_service.get_all_evidence()
        
        # 4. Fill Data
        for row_idx, ev in enumerate(all_evidence, 8):
            conf_val = f"{int(ev.get('confidence') * 100)}%" if isinstance(ev.get('confidence'), (float, int)) and ev.get('confidence') <= 1.0 else f"{ev.get('confidence')}%" if ev.get('confidence') else "85%"
            
            ws.cell(row=row_idx, column=1, value=ev.get("evidence_id")).font = normal_font
            ws.cell(row=row_idx, column=2, value=ev.get("violation_id")).font = normal_font
            ws.cell(row=row_idx, column=3, value=ev.get("vehicle_id")).font = normal_font
            ws.cell(row=row_idx, column=4, value=ev.get("plate_number") or "N/A").font = normal_font
            ws.cell(row=row_idx, column=5, value=ev.get("violation") or "N/A").font = normal_font
            ws.cell(row=row_idx, column=6, value=conf_val).font = normal_font
            ws.cell(row=row_idx, column=7, value=ev.get("camera_id") or "Camera-01").font = normal_font

        # Adjust columns width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(filepath)
        return filepath
